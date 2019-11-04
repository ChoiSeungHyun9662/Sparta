import requests
from bs4 import BeautifulSoup
from openpyxl import load_workbook

#requests는 url에 접근하기 위한 도구. header 메소드를 이용하기 위해 header를 먼저 불러왔다.
headers = {'User-Agent' : 'Mozilla/5.0 (Windows NT 10.0; Win64; x64)AppleWebKit/537.36 (KHTML, like Gecko) Chrome/73.0.3683.86 Safari/537.36'}
data = requests.get('https://movie.naver.com/movie/sdb/rank/rmovie.nhn?sel=pnt&date=20190909',headers=headers)

work_book = load_workbook('naverMovie.xlsx') #어떤 엑셀 파일을 불러올 것인가?
work_sheet = work_book['naverMovie'] #그 파일에서 어떤 시트를 다룰 것인가?

soup = BeautifulSoup(data.text, 'html.parser') #beautifulsoup으로 파싱한 html data를 저장한다.

movies = soup.select('#old_content > table > tbody > tr') #soup에 저장한 html data를 기반으로 select를 사용하여 원하는 자료를 추적한다.
row = 2

for movie in movies:
    name = movie.select_one('td.title > div.tit5 > a')
    score = movie.select_one('td.point')

    if name is not None:
        work_sheet.cell(row=row, column=1, value=row - 1)
        work_sheet.cell(row=row, column=2, value=name.text)
        work_sheet.cell(row=row, column=3, value=float(score.text))
        row += 1

work_book.save('naverMovie.xlsx')
